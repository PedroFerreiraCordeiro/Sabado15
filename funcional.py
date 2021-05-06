# Pedro Gabriel <pedrogabriel.santos@hotmail.com>
# 29/04/2021
#
# Faz busca direto do arquivo txt e retorna um dicionario match
import regex
import re
import os
import win32com.client
from openpyxl import Workbook
import glob


myfile = open('C:\\Users\\pcordeiro\\Desktop\\Projeto abordin\\NOTEPAD\\time_2021_01_22_NF 680 JM SERVIÇOS - TIMBER IX.p-1.jpg.txt','r')
contents= myfile.read()

def find(regular_expression,string):

     return cnpj
string= 'data'
data = re.findall(r'[\d]{1,2}/[\d]{1,2}/[\d]{4}', contents)
print(data)

cnpj = re.findall("[0-9]{2}\.?[0-9]{3}\.?[0-9]{3}\/?[0-9]{4}\-?[0-9]{2}", contents)
print(cnpj)

cep = re.findall("[0-9]{5}\-?[0-9]{3}", contents, flags=re.MULTILINE)
print (cep)

codigoservi = re.findall(r'[0-9]{2}\.?[0-9]{2}\s\-?\s', contents)
print(codigoservi)

nomerazao = re.findall(r'(Nome/Razao Social:|Nome/Razio Social:) \:?(.+)\-?', contents)
print(nomerazao)

endereco = re.findall(r'Endereco: ([a-záéíóúâêôçãõ]+(?: [a-záéíóúâêôçãõ]+)*)', contents, flags=re.IGNORECASE)
print(endereco)

numero = re.findall (r'Numero: [0-9]{3}', contents)
print(numero)

nf = re.findall (r'Nota(.*?)[0-9]{8}', contents, flags=re.IGNORECASE) #[0-9]{3}
print(nf)

bairro = re.findall(r'Bairro: ([a-záéíóúâêôçãõ]+(?: [a-záéíóúâêôçãõ]+)*) ', contents, flags=re.IGNORECASE)
print(bairro)

vt = re.findall(r'(Valor bruto|VALOR TOTAL DA NOTA)\=?(.+)', contents, flags=re.IGNORECASE)
print(vt)

cofins = re.findall(r'(COFINS|Cofins|cofins)(.*\n){2}(.+)', contents, flags=re.MULTILINE) #~R\$:[\t]*(\d{1,3}\.?)+(,\d{2}
print(cofins)
if cofins == []:
    cofins = re.findall(r'(COFINS|Cofins|cofins)(.*\n){2}', contents, flags=re.MULTILINE)
    print(cofins)
else:
    print('texto certo')





match = {
    'CNPJ': cnpj.__str__(),
    'NUMERO NF': '415',
    'DATA EMISSAO': data.__str__(),
    'CODIGO SERVICO': codigoservi.__str__(),
    'VALOR TOTAL': vt.__str__(),
    'NOME': nomerazao.__str__(),
    'INSCRICAO MUNICIPAL': 'INSCRICAO MUNICIPAL',
    'CEP': cep.__str__(),
    'ENDERECO':  endereco.__str__(),
    'NUMERO ENDERECO': numero.__str__(),
    'BAIRRO': bairro.__str__(),
    'COFINS': cofins.__str__(),
    'CSLL': 'CSLL',
    'PISS': 'PISS',
    'IRRF': 'IRRF',
}

myfile.close()

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'CPF/CNPJ'
ws['A2'] = match['CNPJ']
ws['B1'] = "NUMERO NF"
ws['B2'] = match['NUMERO NF']
ws['C1'] = "DATA EMISSAO "
ws['C2'] = match['DATA EMISSAO']
ws['D1'] = "CODIGO SERVICO"
ws['D2'] = match['CODIGO SERVICO']
ws['E1'] = "VALOR TOTAL"
ws['E2'] = match['VALOR TOTAL']
ws['F1'] = "NOME"
ws['F2'] = match['NOME']
ws['G1'] = "INSCRICAO MUNICIPAL"
ws['G2'] = match['INSCRICAO MUNICIPAL']
ws['H1'] = "CEP"
ws['H2'] = match['CEP']
ws['I1'] = "ENDERECO"
ws['I2'] = match['ENDERECO']
ws['J1'] = "NUMERO ENDERECO"
ws['J2'] = match['NUMERO ENDERECO']
ws['K1'] = "BAIRRO "
ws['K2'] = match['BAIRRO']
ws['L1'] = "COFINS"
ws['L2'] = match['COFINS']
ws['M1'] = "CSLL"
ws['M2'] = match['CSLL']
ws['N1'] = "IRRF"
ws['N2'] = match['IRRF']
ws['O1'] = "PISS"
ws['O2'] = match['PISS']

wb.save("C:\\Users\\pcordeiro\\Desktop\\Projeto abordin\\EXCEL\\tese.xlsx")
